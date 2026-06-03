"""Excel output module — dual format support."""

import logging
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

HEADERS = {
    "detailed": [
        "Patient_ID",
        "HPO Term",
        "HPO Code",
        "Present/Absent",
        "Frequency",
        "Onset",
        "Severity",
        "Patient Verbatim",
        "Needs Review",
        "Confidence",
    ],
    "semicolon": ["Patient_ID", "observation_source_value", "excluded_source_value"],
    "purl": ["CASE ID", "HPO TERM", "HPO Code Purl", "Verbatim", "Needs Review", "Confidence"],
}

# Light amber fill for rows the GP should double-check.
_REVIEW_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")


def _review_label(match: dict) -> str:
    """Human-readable review flag for a match row."""
    return "REVIEW" if match.get("needs_review") else ""


def _confidence_cell(match: dict):
    """Confidence as a 0-1 float, or '' when the match carries none."""
    conf = match.get("confidence")
    return conf if conf is not None else ""


# Name of the worksheet that holds the true-path (ancestor-closed) set.
CLOSURE_SHEET = "Ancestor Closure"
CLOSURE_HEADERS = ["Patient_ID", "HPO Term", "HPO Code", "Origin", "Derived From"]

_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")
_TOP_ALIGNMENT = Alignment(vertical="top")


PROVENANCE_SHEET = "Provenance"


def write_excel(
    patient_id: str,
    matches: list[dict],
    output_path: str,
    fmt: str = "semicolon",
    hpo_release: str | None = None,
    propagate_ancestors: bool = False,
    obo_path: str = "data/hpo/hp.obo",
) -> None:
    """Write HPO matches to Excel file.

    Args:
        patient_id: Patient identifier (e.g., "MGA.014").
        matches: List of dicts with hpo_id, hpo_term, patient_verbatim.
        output_path: Path to output Excel file.
        fmt: Output format — "detailed", "semicolon", or "purl".
        hpo_release: Pinned HPO release string (e.g. "hp/releases/2026-02-16").
            Stamped into a Provenance sheet so the workbook is self-describing.
        propagate_ancestors: If True, add an "Ancestor Closure" sheet holding
            the true-path expansion (predicted terms plus all is_a ancestors).
            The primary sheet keeps the leaf-level human view unchanged.
        obo_path: HPO OBO release used to build the is_a graph for propagation.
    """
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    if path.exists():
        wb = load_workbook(path)
        ws = wb[wb.sheetnames[0]]
    else:
        wb = Workbook()
        ws = wb.active
        headers = HEADERS.get(fmt, HEADERS["detailed"])
        ws.append(headers)
        _style_header_row(ws, len(headers))

    if fmt == "purl":
        _write_purl_format(ws, patient_id, matches)
    elif fmt == "semicolon":
        _write_semicolon_format(ws, patient_id, matches)
    else:
        _write_detailed_format(ws, patient_id, matches)

    _auto_fit_columns(ws)
    if hpo_release:
        _stamp_provenance(wb, hpo_release)

    if propagate_ancestors:
        _write_closure_sheet(wb, patient_id, matches, obo_path)

    wb.save(path)
    logger.info("Wrote %d matches for %s to %s (%s format)", len(matches), patient_id, path, fmt)


def _stamp_provenance(wb, hpo_release: str) -> None:
    """Write or refresh a Provenance sheet carrying the pinned HPO release.

    Idempotent: re-running the pipeline against an existing workbook updates the
    release cell rather than appending a second sheet.
    """
    if PROVENANCE_SHEET in wb.sheetnames:
        ws = wb[PROVENANCE_SHEET]
    else:
        ws = wb.create_sheet(PROVENANCE_SHEET)
        ws["A1"] = "Key"
        ws["B1"] = "Value"
        ws["A1"].font = _HEADER_FONT
        ws["B1"].font = _HEADER_FONT
        ws["A1"].fill = _HEADER_FILL
        ws["B1"].fill = _HEADER_FILL
        ws["A2"] = "hpo_release"
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 32
    ws["B2"] = hpo_release


def _write_closure_sheet(wb, patient_id: str, matches: list[dict], obo_path: str) -> None:
    """Append the true-path (ancestor-closed) set to a dedicated sheet.

    The closure sheet is one row per term per patient: leaves plus the
    is_a ancestors they imply. Downstream phenotype tools consume this set.
    """
    # Lazy import: only pay the graph-build cost when propagation is requested.
    from phenoscribe.ontology import propagate_matches

    if CLOSURE_SHEET in wb.sheetnames:
        ws = wb[CLOSURE_SHEET]
    else:
        ws = wb.create_sheet(CLOSURE_SHEET)
        ws.append(CLOSURE_HEADERS)
        _style_header_row(ws, len(CLOSURE_HEADERS))

    closure = propagate_matches(matches, obo_path)
    for row in closure:
        derived = ", ".join(row["derived_from"]) if row["derived_from"] else ""
        ws.append([patient_id, row["hpo_term"], row["hpo_id"], row["origin"], derived])
        r = ws.max_row
        for col in range(1, 5):
            ws.cell(row=r, column=col).alignment = _TOP_ALIGNMENT
        ws.cell(row=r, column=5).alignment = _WRAP_ALIGNMENT

    _auto_fit_columns(ws)
    logger.info(
        "Wrote ancestor closure for %s: %d term(s) on sheet '%s'",
        patient_id, len(closure), CLOSURE_SHEET,
    )


def _style_header_row(ws, num_cols: int) -> None:
    """Apply styling to the header row and freeze it."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"


def _auto_fit_columns(ws) -> None:
    """Auto-fit column widths based on content, with a max width cap."""
    max_width = 60
    for col_idx, col_cells in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), start=1):
        max_len = 0
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        width = min(max_len + 4, max_width)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(width, 12)


def _modifier_cell(m: dict, field: str) -> str:
    """Render a modifier as 'Label (HP:xxxx)' when mapped, else the raw text."""
    hpo_id = m.get(f"{field}_hpo_id")
    hpo_term = m.get(f"{field}_hpo_term")
    if hpo_id and hpo_term:
        return f"{hpo_term} ({hpo_id})"
    return m.get(field, "") or ""


def _write_detailed_format(ws, patient_id: str, matches: list[dict]) -> None:
    """One row per finding with present/absent and modifier columns.

    Negated findings stay in the sheet marked 'Absent' so the excluded
    findings are documented, not silently dropped.
    """
    for m in matches:
        status = "Absent" if m.get("negated", False) else "Present"
        verbatim = m.get("patient_verbatim", "")
        ws.append(
            [
                patient_id,
                m["hpo_term"],
                m["hpo_id"],
                status,
                _modifier_cell(m, "frequency"),
                _modifier_cell(m, "onset"),
                _modifier_cell(m, "severity"),
                verbatim,
                _review_label(m),
                _confidence_cell(m),
            ]
        )
        row = ws.max_row
        for col in range(1, 11):
            ws.cell(row=row, column=col).alignment = _TOP_ALIGNMENT
        ws.cell(row=row, column=8).alignment = _WRAP_ALIGNMENT
        if m.get("needs_review"):
            for col in range(1, 11):
                ws.cell(row=row, column=col).fill = _REVIEW_FILL


def _format_entry(m: dict) -> str:
    """'Name (HP:code) [verbatim]' with modifier annotations appended."""
    entry = f"{m['hpo_term']} ({m['hpo_id']})"
    mods = []
    for field in ("frequency", "onset", "severity"):
        cell = _modifier_cell(m, field)
        if cell:
            mods.append(cell)
    if mods:
        entry += " {" + "; ".join(mods) + "}"
    if m.get("patient_verbatim"):
        entry += f" [{m['patient_verbatim']}]"
    # Mark low-confidence codes inline so they survive the flat cell format.
    if m.get("needs_review"):
        entry += " {REVIEW}"
    return entry


def _write_semicolon_format(ws, patient_id: str, matches: list[dict]) -> None:
    """One row per patient. Present codes and excluded codes in separate columns."""
    present = [m for m in matches if not m.get("negated", False)]
    absent = [m for m in matches if m.get("negated", False)]

    observation = "; ".join(_format_entry(m) for m in present)
    excluded = "; ".join(_format_entry(m) for m in absent)
    ws.append([patient_id, observation, excluded])


def _write_purl_format(ws, patient_id: str, matches: list[dict]) -> None:
    """One row per present HPO term per patient, with PURL-style code links.

    Absent findings are dropped here: this format feeds tools that expect a
    list of phenotypes the patient has.
    """
    for m in matches:
        if m.get("negated", False):
            continue
        purl = _hpo_id_to_purl(m["hpo_id"])
        verbatim = m.get("patient_verbatim", "")
        ws.append(
            [patient_id, m["hpo_term"], purl, verbatim, _review_label(m), _confidence_cell(m)]
        )
        if m.get("needs_review"):
            row = ws.max_row
            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = _REVIEW_FILL


def _hpo_id_to_purl(hpo_id: str) -> str:
    """Convert HP:0002027 to PURL format."""
    obo_id = hpo_id.replace(":", "_")
    return f"http://purl.obolibrary.org/obo/{obo_id}"
