"""Cohort-level prevalence aggregation.

Reads a Phenoscribe results workbook (detailed, semicolon, or PURL
format) and produces the Plovdiv-poster / children's-paper style
output: how many patients had each HPO term, sorted descending.
"""

import csv
import logging
import re
from collections import defaultdict
from pathlib import Path

import openpyxl

logger = logging.getLogger(__name__)

# "Fatigue (HP:0012378) [tired]; Cough (HP:0012735) [coughing]" style.
# The term name is captured non-greedily up to the parenthesised HP code,
# then the bracketed verbatim is optional. Assumes term labels themselves
# do not contain '(' or ';' — HPO labels in the current release don't.
_SEMICOLON_TRIPLET = re.compile(
    r"([^();]+?)\s*\((HP:\d{7})\)(?:\s*\[([^\]]+)\])?"
)
_PURL_TO_HP = re.compile(r"HP_(\d{7})")


def load_patient_codes(path: str) -> dict[str, list[dict]]:
    """Return {patient_id: [{hpo_id, hpo_term, verbatim?}, ...]}.

    Auto-detects detailed / semicolon / PURL format from the headers.
    """
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    result: dict[str, list[dict]] = defaultdict(list)

    if "HPO Code" in headers and "Patient_ID" in headers:
        # Detailed: one row per code
        id_col = headers.index("Patient_ID")
        term_col = headers.index("HPO Term")
        code_col = headers.index("HPO Code")
        verb_col = headers.index("Patient Verbatim") if "Patient Verbatim" in headers else None
        status_col = headers.index("Present/Absent") if "Present/Absent" in headers else None
        for row in ws.iter_rows(min_row=2, values_only=True):
            pid, term, code = row[id_col], row[term_col], row[code_col]
            if not (pid and code and str(code).startswith("HP:")):
                continue
            # Prevalence counts findings the patient has. Skip absent findings.
            if status_col is not None and str(row[status_col]).strip().lower() == "absent":
                continue
            entry = {"hpo_id": str(code), "hpo_term": str(term or "")}
            if verb_col is not None and row[verb_col]:
                entry["verbatim"] = str(row[verb_col])
            result[str(pid)].append(entry)

    elif "HPO Code Purl" in headers:
        id_col = headers.index("CASE ID")
        term_col = headers.index("HPO TERM") if "HPO TERM" in headers else None
        purl_col = headers.index("HPO Code Purl")
        verb_col = headers.index("Verbatim") if "Verbatim" in headers else None
        for row in ws.iter_rows(min_row=2, values_only=True):
            pid, purl = row[id_col], row[purl_col]
            if not (pid and purl):
                continue
            match = _PURL_TO_HP.search(str(purl))
            if not match:
                continue
            entry = {
                "hpo_id": f"HP:{match.group(1)}",
                "hpo_term": str(row[term_col] or "") if term_col is not None else "",
            }
            if verb_col is not None and row[verb_col]:
                entry["verbatim"] = str(row[verb_col])
            result[str(pid)].append(entry)

    else:
        # Semicolon: one row per patient with "Name (HP:code) [verbatim]; ..."
        for row in ws.iter_rows(min_row=2, values_only=True):
            pid = row[0]
            obs = str(row[1]) if row[1] else ""
            if not pid:
                continue
            for term, code, verbatim in _SEMICOLON_TRIPLET.findall(obs):
                entry = {"hpo_id": code, "hpo_term": term.strip()}
                if verbatim:
                    entry["verbatim"] = verbatim.strip()
                result[str(pid)].append(entry)

    return dict(result)


def compute_prevalence(patient_codes: dict[str, list[dict]]) -> list[dict]:
    """Aggregate per-patient codes into per-term prevalence rows.

    Returns rows sorted by n_patients descending, then hpo_id ascending for
    deterministic tie-breaking. Each row: hpo_id, hpo_term, n_patients, pct,
    patient_ids (sorted list).
    """
    n_patients_total = len(patient_codes)
    # term_id -> {"hpo_term": ..., "patient_ids": set()}
    by_term: dict[str, dict] = defaultdict(lambda: {"hpo_term": "", "patient_ids": set()})

    for pid, entries in patient_codes.items():
        for entry in entries:
            hp_id = entry["hpo_id"]
            slot = by_term[hp_id]
            slot["patient_ids"].add(pid)
            incoming = entry.get("hpo_term", "")
            if not slot["hpo_term"] and incoming:
                slot["hpo_term"] = incoming
            elif incoming and slot["hpo_term"] and incoming != slot["hpo_term"]:
                # Rare with Task 1's canonical-name fix in place, but possible
                # if the workbook mixes older outputs or manual edits.
                logger.debug(
                    "term_label_disagreement: id=%s kept=%r dropped=%r",
                    hp_id, slot["hpo_term"], incoming,
                )

    rows = [
        {
            "hpo_id": hp_id,
            "hpo_term": slot["hpo_term"],
            "n_patients": len(slot["patient_ids"]),
            "pct": (len(slot["patient_ids"]) / n_patients_total * 100) if n_patients_total else 0.0,
            "patient_ids": sorted(slot["patient_ids"]),
        }
        for hp_id, slot in by_term.items()
    ]
    rows.sort(key=lambda r: (-r["n_patients"], r["hpo_id"]))
    return rows


def write_prevalence_csv(rows: list[dict], path: str) -> None:
    """Write prevalence rows to CSV. Patient IDs are joined with '|' to fit one cell."""
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["hpo_id", "hpo_term", "n_patients", "pct", "patient_ids"])
        for r in rows:
            writer.writerow([
                r["hpo_id"],
                r["hpo_term"],
                r["n_patients"],
                f"{r['pct']:.1f}",
                "|".join(r["patient_ids"]),
            ])
    logger.info("Wrote prevalence CSV (%d terms) to %s", len(rows), path)


def write_prevalence_chart(
    rows: list[dict],
    path: str,
    top_n: int = 20,
    n_patients: int | None = None,
) -> None:
    """Write a horizontal bar chart of the top N terms by patient count.

    Layout matches the Plovdiv-poster style: terms on the y-axis, count on the x-axis,
    longest bars at the top. `n_patients` is the cohort size that the bars are
    expressed against; if omitted the title falls back to the count of distinct terms.
    """
    if not rows:
        logger.warning("No prevalence rows to chart; skipping %s", path)
        return

    # Lazy import — matplotlib is heavy and only needed when a chart is requested.
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    top = rows[:top_n]
    # barh draws the first row at the bottom; reverse so the longest is at the top.
    labels = [f"{r['hpo_term']} ({r['hpo_id']})" for r in reversed(top)]
    counts = [r["n_patients"] for r in reversed(top)]

    fig_height = max(3.0, 0.35 * len(top) + 1.5)
    fig, ax = plt.subplots(figsize=(10, fig_height))
    ax.barh(labels, counts, color="#4472C4")
    cohort_suffix = f", N={n_patients} patients" if n_patients is not None else ""
    ax.set_xlabel("Patients")
    ax.set_title(
        f"HPO term prevalence (top {len(top)} of {len(rows)} terms{cohort_suffix})"
    )
    ax.xaxis.get_major_locator().set_params(integer=True)
    for i, c in enumerate(counts):
        ax.text(c, i, f" {c}", va="center", fontsize=9)
    fig.tight_layout()

    Path(path).parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(path, dpi=120)
    plt.close(fig)
    logger.info("Wrote prevalence chart (%d terms shown) to %s", len(top), path)
