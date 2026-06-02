"""Tests that needs_review and confidence reach the output workbook.

Robinson's complaint: a code the model could not justify must not look like one
it endorsed. The output has to carry the review flag in every format.
"""

import openpyxl

from phenoscribe.output import write_excel


MATCHES = [
    {
        "hpo_id": "HP:0012378",
        "hpo_term": "Fatigue",
        "patient_verbatim": "fatigué",
        "confidence": 0.92,
        "needs_review": False,
    },
    {
        "hpo_id": "HP:0002027",
        "hpo_term": "Abdominal pain",
        "patient_verbatim": "mal au ventre",
        "confidence": 0.0,
        "needs_review": True,
    },
]


def _read(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    return headers, rows


def test_detailed_format_has_review_and_confidence_columns(tmp_path):
    path = tmp_path / "out.xlsx"
    write_excel("MGA.014", MATCHES, str(path), fmt="detailed")

    headers, rows = _read(path)
    assert "Needs Review" in headers
    assert "Confidence" in headers

    review_col = headers.index("Needs Review")
    conf_col = headers.index("Confidence")

    # First row endorsed, second flagged.
    assert rows[0][review_col] in ("", None)
    assert rows[0][conf_col] == 0.92
    assert rows[1][review_col] == "REVIEW"


def test_purl_format_has_review_column(tmp_path):
    path = tmp_path / "out.xlsx"
    write_excel("MGA.014", MATCHES, str(path), fmt="purl")

    headers, rows = _read(path)
    assert "Needs Review" in headers
    review_col = headers.index("Needs Review")
    assert rows[1][review_col] == "REVIEW"


def test_semicolon_format_marks_review_inline(tmp_path):
    path = tmp_path / "out.xlsx"
    write_excel("MGA.014", MATCHES, str(path), fmt="semicolon")

    _, rows = _read(path)
    observation = rows[0][1]
    # The endorsed code has no marker; the flagged one does.
    assert "Fatigue (HP:0012378)" in observation
    assert "{REVIEW}" in observation
    # Marker attaches to the flagged code, not the endorsed one.
    assert "Fatigue (HP:0012378) [fatigué] {REVIEW}" not in observation
    assert "{REVIEW}" in observation.split("Abdominal pain")[1]


def test_matches_without_flags_still_write(tmp_path):
    """Older match dicts that predate the confidence fields must not crash."""
    path = tmp_path / "out.xlsx"
    legacy = [{"hpo_id": "HP:0012378", "hpo_term": "Fatigue", "patient_verbatim": "x"}]
    write_excel("P1", legacy, str(path), fmt="detailed")

    headers, rows = _read(path)
    review_col = headers.index("Needs Review")
    conf_col = headers.index("Confidence")
    assert rows[0][review_col] in ("", None)
    assert rows[0][conf_col] in ("", None)
