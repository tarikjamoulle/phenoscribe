"""Tests for the pinned HPO release: version guard, obsolete resolution, provenance.

Answers Robinson issue #1 — the obo on disk, the index and config must agree on
one release, and an obsolete ground-truth code must resolve to its replacement
instead of silently dropping out of search.
"""

from pathlib import Path

import pytest

from phenoscribe.hpo_index import (
    HpoVersionMismatch,
    build_obsolete_map,
    check_obo_version,
    parse_obo,
    read_obo_version,
    resolve_obsolete,
)
from phenoscribe.output import write_excel
from phenoscribe.validate import load_codes_from_excel

# The real release shipped in the repo.
REAL_OBO = "/Users/tarikjamoulle/projects/hpo_identifier/data/hpo/hp.obo"
PINNED_RELEASE = "hp/releases/2026-02-16"

# A tiny obo with a known obsolete-with-replaced_by term, an obsolete
# consider-only term, and a merged term (alt_id). Lets the resolution logic be
# checked without parsing the full 19k-term release.
MINI_OBO = """\
format-version: 1.2
data-version: hp/releases/2099-01-01

[Term]
id: HP:0000001
name: Live term
synonym: "alive" EXACT []
def: "A non-obsolete term." [HPO:probinson]

[Term]
id: HP:0000002
name: Replacement target

[Term]
id: HP:0000003
name: obsolete Old term
is_obsolete: true
replaced_by: HP:0000002

[Term]
id: HP:0000004
name: obsolete Consider-only term
is_obsolete: true
consider: HP:0000001
consider: HP:0000002

[Term]
id: HP:0000005
name: Merged survivor
alt_id: HP:0000099

[Typedef]
id: part_of
name: part of
"""


@pytest.fixture
def mini_obo(tmp_path) -> str:
    p = tmp_path / "mini.obo"
    p.write_text(MINI_OBO)
    return str(p)


# --- version header reading ---------------------------------------------------


def test_read_obo_version_real_release():
    assert read_obo_version(REAL_OBO) == PINNED_RELEASE


def test_read_obo_version_missing_file_raises(tmp_path):
    with pytest.raises(FileNotFoundError):
        read_obo_version(str(tmp_path / "nope.obo"))


def test_read_obo_version_no_header_raises(tmp_path):
    p = tmp_path / "headerless.obo"
    p.write_text("[Term]\nid: HP:0000001\nname: x\n")
    with pytest.raises(ValueError):
        read_obo_version(str(p))


# --- the build/startup guard --------------------------------------------------


def test_check_obo_version_passes_on_match():
    assert check_obo_version(REAL_OBO, PINNED_RELEASE) == PINNED_RELEASE


def test_check_obo_version_raises_on_mismatch():
    # This is exactly Robinson's scenario: Dockerfile/config pin 2025-05-06 but
    # the obo on disk is 2026-02-16. The guard must fail loudly.
    with pytest.raises(HpoVersionMismatch) as exc:
        check_obo_version(REAL_OBO, "hp/releases/2025-05-06")
    assert "2025-05-06" in str(exc.value)
    assert PINNED_RELEASE in str(exc.value)


# --- obsolete / merged id resolution ------------------------------------------


def test_parse_obo_drops_obsolete_terms(mini_obo):
    ids = {t["id"] for t in parse_obo(mini_obo)}
    assert "HP:0000001" in ids
    assert "HP:0000003" not in ids  # obsolete
    assert "HP:0000004" not in ids  # obsolete


def test_resolve_replaced_by(mini_obo):
    m = build_obsolete_map(mini_obo)
    assert resolve_obsolete("HP:0000003", m) == "HP:0000002"


def test_resolve_merged_alt_id(mini_obo):
    m = build_obsolete_map(mini_obo)
    assert resolve_obsolete("HP:0000099", m) == "HP:0000005"


def test_consider_only_term_not_auto_resolved(mini_obo):
    # consider: needs human review, so we never rewrite it automatically.
    m = build_obsolete_map(mini_obo)
    assert resolve_obsolete("HP:0000004", m) == "HP:0000004"


def test_resolve_passthrough_for_live_id(mini_obo):
    m = build_obsolete_map(mini_obo)
    assert resolve_obsolete("HP:0000001", m) == "HP:0000001"


def test_resolve_real_obsolete_with_replaced_by():
    # HP:0000057 "obsolete Clitoromegaly" -> replaced_by HP:0008665 in the
    # 2026-02-16 release. A real obsoleted term with a replacement.
    m = build_obsolete_map(REAL_OBO)
    assert resolve_obsolete("HP:0000057", m) == "HP:0008665"


def test_resolve_real_ground_truth_obsolete_code():
    # HP:0001322 appears in Marc Jamoulle's ground truth and is obsolete in
    # 2026-02-16, replaced_by HP:0006872. Without resolution it would never
    # match a current prediction.
    m = build_obsolete_map(REAL_OBO)
    assert resolve_obsolete("HP:0001322", m) == "HP:0006872"


# --- ground-truth loading resolves obsolete codes -----------------------------


def test_load_codes_resolves_obsolete(tmp_path):
    out = tmp_path / "gt.xlsx"
    # Workbook carrying the obsolete code HP:0000003.
    write_excel(
        "PAT.001",
        [{"hpo_id": "HP:0000003", "hpo_term": "Old term", "patient_verbatim": "x"}],
        str(out),
        fmt="detailed",
    )
    # Build the map from the mini obo.
    mini = tmp_path / "mini.obo"
    mini.write_text(MINI_OBO)
    obsolete_map = build_obsolete_map(str(mini))

    raw = load_codes_from_excel(str(out))
    resolved = load_codes_from_excel(str(out), obsolete_map)
    assert raw["PAT.001"] == {"HP:0000003"}
    assert resolved["PAT.001"] == {"HP:0000002"}


# --- provenance stamp ---------------------------------------------------------


def test_workbook_carries_release_string(tmp_path):
    import openpyxl

    out = tmp_path / "results.xlsx"
    write_excel(
        "PAT.001",
        [{"hpo_id": "HP:0000001", "hpo_term": "Live term", "patient_verbatim": "v"}],
        str(out),
        fmt="detailed",
        hpo_release=PINNED_RELEASE,
    )
    wb = openpyxl.load_workbook(str(out))
    assert "Provenance" in wb.sheetnames
    ws = wb["Provenance"]
    assert ws["A2"].value == "hpo_release"
    assert ws["B2"].value == PINNED_RELEASE
    # The data sheet must still be the active/first sheet so loaders find codes.
    assert wb.sheetnames[0] != "Provenance"


def test_provenance_stamp_is_idempotent(tmp_path):
    import openpyxl

    out = tmp_path / "results.xlsx"
    for _ in range(2):
        write_excel(
            "PAT.001",
            [{"hpo_id": "HP:0000001", "hpo_term": "Live term", "patient_verbatim": "v"}],
            str(out),
            fmt="detailed",
            hpo_release=PINNED_RELEASE,
        )
    wb = openpyxl.load_workbook(str(out))
    assert wb.sheetnames.count("Provenance") == 1
    assert wb["Provenance"]["B2"].value == PINNED_RELEASE
