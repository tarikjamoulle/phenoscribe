"""Tests for the cohort aggregation module."""

from pathlib import Path

import openpyxl
import pytest

from phenoscribe.aggregate import (
    compute_prevalence,
    load_patient_codes,
    write_prevalence_chart,
    write_prevalence_csv,
)


def _make_detailed_xlsx(tmp_path: Path, rows: list[tuple]) -> Path:
    path = tmp_path / "results.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Patient_ID", "HPO Term", "HPO Code", "Patient Verbatim"])
    for row in rows:
        ws.append(list(row))
    wb.save(path)
    return path


def _make_purl_xlsx(tmp_path: Path, rows: list[tuple]) -> Path:
    path = tmp_path / "results_purl.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["CASE ID", "HPO TERM", "HPO Code Purl", "Verbatim"])
    for pid, term, hp_id, verbatim in rows:
        purl = f"http://purl.obolibrary.org/obo/HP_{hp_id.split(':')[1]}"
        ws.append([pid, term, purl, verbatim])
    wb.save(path)
    return path


def _make_semicolon_xlsx(tmp_path: Path, rows: list[tuple[str, str]]) -> Path:
    path = tmp_path / "results_semi.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Patient_ID", "observation_source_value"])
    for pid, obs in rows:
        ws.append([pid, obs])
    wb.save(path)
    return path


def test_load_detailed_format(tmp_path):
    path = _make_detailed_xlsx(
        tmp_path,
        [
            ("P1", "Fatigue", "HP:0012378", "tired"),
            ("P1", "Cough", "HP:0012735", "coughing"),
            ("P2", "Fatigue", "HP:0012378", "exhausted"),
        ],
    )
    out = load_patient_codes(str(path))
    assert set(out.keys()) == {"P1", "P2"}
    assert len(out["P1"]) == 2
    assert {e["hpo_id"] for e in out["P1"]} == {"HP:0012378", "HP:0012735"}
    assert out["P1"][0]["verbatim"] == "tired"


def test_load_purl_format(tmp_path):
    path = _make_purl_xlsx(
        tmp_path,
        [
            ("P1", "Fatigue", "HP:0012378", "tired"),
            ("P2", "Fatigue", "HP:0012378", "exhausted"),
        ],
    )
    out = load_patient_codes(str(path))
    assert out["P1"][0]["hpo_id"] == "HP:0012378"
    assert out["P1"][0]["hpo_term"] == "Fatigue"


def test_load_semicolon_format(tmp_path):
    path = _make_semicolon_xlsx(
        tmp_path,
        [
            ("P1", "Fatigue (HP:0012378) [tired]; Cough (HP:0012735) [coughing]"),
            ("P2", "Fatigue (HP:0012378) [exhausted]"),
        ],
    )
    out = load_patient_codes(str(path))
    assert {e["hpo_id"] for e in out["P1"]} == {"HP:0012378", "HP:0012735"}
    assert any(e["hpo_term"] == "Fatigue" for e in out["P1"])


def test_prevalence_shared_term_tops_the_list():
    patient_codes = {
        "P1": [{"hpo_id": "HP:0012378", "hpo_term": "Fatigue"}, {"hpo_id": "HP:0012735", "hpo_term": "Cough"}],
        "P2": [{"hpo_id": "HP:0012378", "hpo_term": "Fatigue"}, {"hpo_id": "HP:0002027", "hpo_term": "Abdominal pain"}],
        "P3": [{"hpo_id": "HP:0012378", "hpo_term": "Fatigue"}],
    }
    rows = compute_prevalence(patient_codes)
    assert rows[0]["hpo_id"] == "HP:0012378"
    assert rows[0]["n_patients"] == 3
    assert rows[0]["pct"] == pytest.approx(100.0)
    assert rows[0]["patient_ids"] == ["P1", "P2", "P3"]


def test_prevalence_same_patient_listed_once_per_term():
    # Same patient mentioning the same term twice (e.g. via two utterances)
    # should still count as one patient for that term.
    patient_codes = {
        "P1": [
            {"hpo_id": "HP:0012378", "hpo_term": "Fatigue"},
            {"hpo_id": "HP:0012378", "hpo_term": "Fatigue"},
        ],
    }
    rows = compute_prevalence(patient_codes)
    assert rows[0]["n_patients"] == 1


def test_prevalence_empty_input():
    assert compute_prevalence({}) == []


def test_csv_round_trip(tmp_path):
    rows = [
        {"hpo_id": "HP:0012378", "hpo_term": "Fatigue", "n_patients": 3, "pct": 100.0, "patient_ids": ["P1", "P2", "P3"]},
    ]
    out = tmp_path / "prev.csv"
    write_prevalence_csv(rows, str(out))
    text = out.read_text()
    assert "hpo_id,hpo_term,n_patients,pct,patient_ids" in text
    assert "HP:0012378,Fatigue,3,100.0,P1|P2|P3" in text


def test_chart_writes_a_valid_png(tmp_path):
    rows = [
        {"hpo_id": "HP:0012378", "hpo_term": "Fatigue", "n_patients": 3, "pct": 100.0, "patient_ids": ["P1", "P2", "P3"]},
        {"hpo_id": "HP:0012735", "hpo_term": "Cough", "n_patients": 1, "pct": 33.3, "patient_ids": ["P1"]},
    ]
    out = tmp_path / "prev.png"
    write_prevalence_chart(rows, str(out), top_n=20)
    assert out.exists()
    # PNG magic bytes
    assert out.read_bytes()[:8] == b"\x89PNG\r\n\x1a\n"


def test_chart_skipped_for_empty_rows(tmp_path):
    out = tmp_path / "prev.png"
    write_prevalence_chart([], str(out), top_n=20)
    assert not out.exists()
