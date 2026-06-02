"""Tests for the HPO true-path (ancestor propagation) rule.

Robinson issue #3 / Robinson Test Q7: HPO is a DAG and downstream tools
assume the annotation set is ancestor-closed. These tests pin the closure
against the real ontology release so the true-path expansion is correct.
"""

import openpyxl
import pytest

from phenoscribe.ontology import (
    HPO_ROOT,
    PHENOTYPIC_ABNORMALITY,
    load_hpo_graph,
    propagate_matches,
)
from phenoscribe.output import CLOSURE_SHEET, write_excel

OBO_PATH = "/Users/tarikjamoulle/projects/hpo_identifier/data/hpo/hp.obo"

# Episodic ataxia is a child of Ataxia, deep under Phenotypic abnormality.
EPISODIC_ATAXIA = "HP:0002131"
ATAXIA = "HP:0001251"


@pytest.fixture(scope="module")
def hpo():
    return load_hpo_graph(OBO_PATH)


def test_episodic_ataxia_closure_reaches_ataxia_and_root(hpo):
    ancestors = hpo.ancestors(EPISODIC_ATAXIA)
    assert ATAXIA in ancestors
    assert PHENOTYPIC_ABNORMALITY in ancestors
    assert HPO_ROOT in ancestors
    # The term itself is not in its own ancestor set unless asked for.
    assert EPISODIC_ATAXIA not in ancestors


def test_include_source_adds_the_term_itself(hpo):
    with_self = hpo.ancestors(EPISODIC_ATAXIA, include_source=True)
    assert EPISODIC_ATAXIA in with_self
    assert ATAXIA in with_self


def test_closure_set_is_ancestor_closed(hpo):
    closure = hpo.closure([EPISODIC_ATAXIA])
    assert EPISODIC_ATAXIA in closure
    assert ATAXIA in closure
    assert PHENOTYPIC_ABNORMALITY in closure
    # An ancestor-closed set contains every ancestor of every member.
    for member in list(closure):
        assert hpo.ancestors(member).issubset(closure)


def test_unknown_id_does_not_crash(hpo):
    assert hpo.ancestors("HP:9999999") == set()
    assert hpo.ancestors("HP:9999999", include_source=True) == {"HP:9999999"}


def test_propagate_matches_keeps_leaves_and_adds_ancestors():
    matches = [
        {"hpo_id": EPISODIC_ATAXIA, "hpo_term": "Episodic ataxia",
         "patient_verbatim": "comes and goes"},
    ]
    rows = propagate_matches(matches, OBO_PATH)
    by_id = {r["hpo_id"]: r for r in rows}

    # Leaf preserved and flagged.
    assert by_id[EPISODIC_ATAXIA]["origin"] == "leaf"
    # Ancestors present, flagged, and traceable to the leaf.
    assert by_id[ATAXIA]["origin"] == "ancestor"
    assert EPISODIC_ATAXIA in by_id[ATAXIA]["derived_from"]
    assert by_id[ATAXIA]["hpo_term"] == "Ataxia"
    assert PHENOTYPIC_ABNORMALITY in by_id


def test_explicit_leaf_stays_a_leaf_even_if_it_is_an_ancestor():
    # If both a term and its ancestor are predicted directly, the ancestor is
    # an explicit leaf, not a derived row.
    matches = [
        {"hpo_id": EPISODIC_ATAXIA, "hpo_term": "Episodic ataxia"},
        {"hpo_id": ATAXIA, "hpo_term": "Ataxia"},
    ]
    rows = propagate_matches(matches, OBO_PATH)
    ataxia_rows = [r for r in rows if r["hpo_id"] == ATAXIA]
    assert len(ataxia_rows) == 1
    assert ataxia_rows[0]["origin"] == "leaf"


def test_write_excel_no_propagation_leaves_single_sheet(tmp_path):
    out = tmp_path / "results.xlsx"
    matches = [{"hpo_id": EPISODIC_ATAXIA, "hpo_term": "Episodic ataxia",
                "patient_verbatim": "v"}]
    write_excel("P1", matches, str(out), fmt="detailed", propagate_ancestors=False)

    wb = openpyxl.load_workbook(out)
    assert CLOSURE_SHEET not in wb.sheetnames
    ws = wb[wb.sheetnames[0]]
    codes = [row[2] for row in ws.iter_rows(min_row=2, values_only=True)]
    assert codes == [EPISODIC_ATAXIA]


def test_write_excel_with_propagation_adds_closure_sheet(tmp_path):
    out = tmp_path / "results.xlsx"
    matches = [{"hpo_id": EPISODIC_ATAXIA, "hpo_term": "Episodic ataxia",
                "patient_verbatim": "v"}]
    write_excel(
        "P1", matches, str(out), fmt="detailed",
        propagate_ancestors=True, obo_path=OBO_PATH,
    )

    wb = openpyxl.load_workbook(out)
    # Primary (leaf) sheet unchanged: still one data row, the leaf only.
    primary = wb[wb.sheetnames[0]]
    primary_codes = [row[2] for row in primary.iter_rows(min_row=2, values_only=True)]
    assert primary_codes == [EPISODIC_ATAXIA]

    # Closure sheet present and ancestor-closed.
    assert CLOSURE_SHEET in wb.sheetnames
    closure = wb[CLOSURE_SHEET]
    closure_codes = {row[2] for row in closure.iter_rows(min_row=2, values_only=True)}
    assert EPISODIC_ATAXIA in closure_codes
    assert ATAXIA in closure_codes
    assert PHENOTYPIC_ABNORMALITY in closure_codes
