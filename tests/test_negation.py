"""Negation handling end to end through match_hpo and output.

Robinson issue #6 + Robinson Test Q5/Q6: a denied finding must not become a
present code. We feed stub extraction records (no live LLM) and assert:
- negated findings are excluded from the present code set
- negated findings are still reported as absent, not dropped
- frequency/onset/severity are carried and mapped to HPO subontology leaves
- the detailed Excel marks absent rows and prevalence skips them
"""

import os

import openpyxl
import pytest

import phenoscribe.match_hpo as match_hpo
from phenoscribe.match_hpo import match_hpo as run_match, present_codes, absent_codes
from phenoscribe.output import write_excel
from phenoscribe.aggregate import load_patient_codes, compute_prevalence


# Map each clinical term to the HPO candidate the stubbed search returns.
_FAKE_INDEX = {
    "fever": {"hpo_id": "HP:0001945", "name": "Fever"},
    "joint pain": {"hpo_id": "HP:0002829", "name": "Arthralgia"},
}


def _stub_pipeline(monkeypatch):
    def fake_search(clinical_term, k=5, chroma_path="", **kwargs):
        hit = _FAKE_INDEX.get(clinical_term.lower())
        return [hit] if hit else []

    def fake_llm(system_prompt, user_prompt, **kwargs):
        # Pick whichever candidate id appears in the prompt.
        for hit in _FAKE_INDEX.values():
            if hit["name"] in user_prompt:
                return f'{{"hpo_id": "{hit["hpo_id"]}", "hpo_term": "{hit["name"]}"}}'
        return "{}"

    monkeypatch.setattr(match_hpo, "search_hpo", fake_search)
    monkeypatch.setattr(match_hpo, "llm_call", fake_llm)


# "je n'ai pas de fievre, mais des douleurs aux articulations"
SYMPTOMS = [
    {
        "patient_verbatim": "je n'ai pas de fievre",
        "clinical_term": "fever",
        "negated": True,
        "frequency": "",
        "onset": "",
        "severity": "",
        "context": "",
    },
    {
        "patient_verbatim": "des douleurs aux articulations",
        "clinical_term": "joint pain",
        "negated": False,
        "frequency": "frequent",
        "onset": "adult",
        "severity": "severe",
        "context": "",
    },
]


def test_negated_finding_excluded_from_present_codes(monkeypatch):
    _stub_pipeline(monkeypatch)

    matches = run_match(SYMPTOMS, chroma_path="")

    present = present_codes(matches)
    absent = absent_codes(matches)

    present_ids = {m["hpo_id"] for m in present}
    absent_ids = {m["hpo_id"] for m in absent}

    # Joint pain (Arthralgia) is present; fever is NOT emitted as present.
    assert present_ids == {"HP:0002829"}
    assert "HP:0001945" not in present_ids
    # Fever is still reported, marked absent — denied findings are not dropped.
    assert absent_ids == {"HP:0001945"}


def test_modifiers_mapped_to_hpo_leaves(monkeypatch):
    _stub_pipeline(monkeypatch)

    matches = run_match(SYMPTOMS, chroma_path="")
    arthralgia = next(m for m in matches if m["hpo_id"] == "HP:0002829")

    assert arthralgia["frequency_hpo_id"] == "HP:0040282"
    assert arthralgia["onset_hpo_id"] == "HP:0003581"
    assert arthralgia["severity_hpo_id"] == "HP:0012828"


def test_detailed_excel_marks_absent_and_prevalence_skips_it(monkeypatch, tmp_path):
    _stub_pipeline(monkeypatch)
    matches = run_match(SYMPTOMS, chroma_path="")

    out = tmp_path / "out.xlsx"
    write_excel("P001", matches, str(out), fmt="detailed")

    codes = load_patient_codes(str(out))
    # The absent fever row must not count toward the patient's phenotypes.
    patient_ids = codes["P001"]
    assert {e["hpo_id"] for e in patient_ids} == {"HP:0002829"}

    rows = compute_prevalence(codes)
    by_id = {r["hpo_id"]: r for r in rows}
    assert "HP:0001945" not in by_id
    assert by_id["HP:0002829"]["n_patients"] == 1


def test_semicolon_excel_splits_present_and_excluded(monkeypatch, tmp_path):
    _stub_pipeline(monkeypatch)
    matches = run_match(SYMPTOMS, chroma_path="")

    out = tmp_path / "semi.xlsx"
    write_excel("P001", matches, str(out), fmt="semicolon")

    ws = openpyxl.load_workbook(out).active
    headers = [c.value for c in ws[1]]
    row = [c.value for c in ws[2]]
    present = row[headers.index("observation_source_value")] or ""
    excluded = row[headers.index("excluded_source_value")] or ""

    assert "HP:0002829" in present
    assert "HP:0001945" not in present
    assert "HP:0001945" in excluded


@pytest.mark.skipif(
    not os.environ.get("ANTHROPIC_API_KEY"),
    reason="live LLM smoke test; needs ANTHROPIC_API_KEY",
)
def test_live_negation_smoke():
    """One live call: 'je n'ai pas de fievre, mais des douleurs aux articulations'.

    Asserts fever is negated/absent and joint pain is present. Skipped unless
    ANTHROPIC_API_KEY is set so the deterministic suite stays offline.
    """
    from phenoscribe.extract_symptoms import extract_symptoms

    text = "je n'ai pas de fievre, mais des douleurs aux articulations"
    res = extract_symptoms(text, provider="anthropic", model="claude-sonnet-4-6")

    by_term = {r["clinical_term"].lower(): r for r in res}
    fever = next((v for k, v in by_term.items() if "fever" in k), None)
    pain = next((v for k, v in by_term.items() if "pain" in k or "arthralgia" in k), None)

    assert fever is not None and fever["negated"] is True
    assert pain is not None and pain["negated"] is False
